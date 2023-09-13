from openpyxl import Workbook
from openpyxl.styles import Font

def set_data(sheet, header, data_list):
    for i, head in enumerate(header, 1):
        cell = sheet.cell(row=1, column=i, value=head)
        cell.font = Font(bold=True)
    for i, data in enumerate(data_list, 2):
        for j, item in enumerate(data, 1):
            sheet.cell(row=i, column=j, value=item)

def search_info(sheet, code, col_idx, return_col_idx):
    for i in range(2, 5):
        if sheet.cell(row=i, column=col_idx).value == code:
            return sheet.cell(row=i, column=return_col_idx).value

wb = Workbook()
wb.active.title = "Language"
wb.create_sheet(title="Capital")

header1 = ["State", "Language", "Code"]
data1 = [("Karnataka", "Kannada", 'KA'), ("Telangana", "Telugu", 'TS'), ("Tamil Nadu", "Tamil", 'TN')]

header2 = ["State", "Capital", "Code"]
data2 = [("Karnataka", "Bengaluru", 'KA'), ("Telangana", "Hyderabad", 'TS'), ("Tamil Nadu", "Chennai", 'TN')]

set_data(wb["Language"], header1, data1)
set_data(wb["Capital"], header2, data2)
wb.save("demo.xlsx")

srch_code = input("Enter state code for finding capital: ")
capital = search_info(wb["Capital"], srch_code, 3, 2)
if capital:
    print(f"Corresponding capital for code {srch_code} is {capital}")

srch_code = input("Enter state code for finding language: ")
language = search_info(wb["Language"], srch_code, 3, 2)
if language:
    print(f"Corresponding language for code {srch_code} is {language}")

wb.close()
